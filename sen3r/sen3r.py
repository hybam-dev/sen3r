import os
import sys
import time
import pandas as pd
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

from sen3r.commons import Utils, DefaultDicts, Footprinter
from sen3r.nc_engine import NcEngine, ParallelBandExtract
from sen3r.tsgen import TsGenerator


if sys.version_info >= (3, 8):
    from importlib import metadata
else:
    import importlib_metadata as metadata

dd = DefaultDicts()


class Core:
    """
    Core methods to build the intermediate files from input Sentinel-3 NetCDF4 images.
    """

    def __init__(self, input_args: dict):
        self.arguments = input_args
        self.INPUT_DIR = self.arguments['input']
        self.OUTPUT_DIR = self.arguments['out']
        Path(self.OUTPUT_DIR).mkdir(parents=True, exist_ok=True)  # Assure the existence of an output folder for the LOG
        self.ROI = self.arguments['roi']
        self.RNAME = os.path.basename(self.ROI.split('.')[0])  # Take the ROI without the file extension.
        self.product = self.arguments['product']
        self.CSV_N1 = os.path.join(self.OUTPUT_DIR, 'CSV_N1')
        self.REP = os.path.join(self.OUTPUT_DIR, 'RDATA')
        self.INSTANCE_TIME_TAG = datetime.now().strftime('%Y%m%dT%H%M%S')
        self.arguments['logfile'] = os.path.join(self.arguments['out'], 'sen3r_' + self.INSTANCE_TIME_TAG + '.log')
        self.log = Utils.create_log_handler(self.arguments['logfile'])
        self.IMG_DIR = os.path.join(self.OUTPUT_DIR, 'images')
        # Section 5 for single source of truth for the version number:
        # https://packaging.python.org/guides/single-sourcing-package-version/#single-sourcing-the-version
        self.VERSION = metadata.version('sen3r')  # TODO: May be outdated depending on the environment installed version
        self.vertices = None  # Further declaration may happen inside build_intermediary_files
        self.sorted_file_list = None  # Declaration may happen inside build_intermediary_files

    @staticmethod
    def build_list_from_subset(input_directory_path):
        """
        Creates a python list containing the Posixpath from all the files inside the directory and sort them by date.
        """
        # convert input string to Posix
        in_path_obj = Path(input_directory_path)
        # get only the '20160425T134227' from the file name and use it to sort the list by date
        sorted_output_files = sorted(os.listdir(in_path_obj), key=lambda s: s.split('____')[1].split('_')[0])
        sorted_output_files_fullpath = [os.path.join(in_path_obj, img) for img in sorted_output_files]

        return sorted_output_files_fullpath

    def get_s3_data(self, wfr_img_folder, vertices=None, roi_file=None, rgb=True, parallel=True):
        """
        Given a vector and a S3_OL2_WFR image, extract the NC data inside the vector.
        """
        img_data = {}
        img = wfr_img_folder
        process_flag = True

        # Test if there is a footprint.shp file inside the image folder
        footprint = Path(img) / 'footprint.shp'
        if footprint.is_file():
            self.log.info('footprint.shp found inside image folder.')
            # Test if the footprint.shp touches the user ROI
            if not Footprinter.touch_test(footprint, roi_file):
                # If touch_test returns False the ROI don't touch and the image can be skipped
                process_flag = False
            else:
                self.log.info('User ROI touches footprint.shp, processing will continue.')

        if process_flag:
            # Class instance of NcEngine containing information about all the bands.
            nce = NcEngine(input_nc_folder=img, parent_log=self.log)

            # Convert the input ROI LAT/LON vertices to X,Y coordinates based on the geo_coordinates.nc file
            img_data['xy_vert'], img_data['ll_vert'] = nce.latlon_2_xy_poly(poly_path=vertices)

            # II) Use the poly to generate an extraction mask:
            img_data['mask'], img_data['cc'], img_data['rr'] = nce.get_raster_mask(xy_vertices=img_data['xy_vert'])

            # III) Get the dictionary of available bands based on the product:
            if self.product and self.product.lower() == 'wfr':
                img_data['bdict'] = dd.wfr_files
            elif self.product and self.product.lower() == 'syn':
                img_data['bdict'] = dd.syn_files
            else:
                self.log.info(f'Invalid product: {self.product.upper()}.')
                sys.exit(1)

            img_data['g_lon'] = nce.g_lon
            img_data['g_lat'] = nce.g_lat
            img_data['OAA'] = nce.OAA
            img_data['OZA'] = nce.OZA
            img_data['SAA'] = nce.SAA
            img_data['SZA'] = nce.SZA
            img_data['nc_file'] = nce.nc_folder

            # IV) Extract the data from the NetCDF using the mask
            pbe = ParallelBandExtract()
            df = pbe.nc_2_df(rr=img_data['rr'], cc=img_data['cc'],
                             lon=img_data['g_lon'],
                             lat=img_data['g_lat'],
                             oaa=img_data['OAA'],
                             oza=img_data['OZA'],
                             saa=img_data['SAA'],
                             sza=img_data['SZA'],
                             nc_folder=img_data['nc_file'],
                             wfr_files_p=dd.wfr_files_p,
                             parent_log=self.arguments['logfile'])

            if self.product.lower() == 'wfr':
                df = df.rename(columns=dd.wfr_vld_names)

            # TODO: check necessity of renaming SYNERGY colnames.
            # if self.product.lower() == 'syn':
            #     df = df.rename(columns=self.syn_vld_names)

            if len(df) == 0:
                self.log.info('EMPTY DATAFRAME WARNING! Unable to find valid pixels in file.')

            img_data['colors'] = {}
            img_data['img'] = None
            if rgb:
                img_data['colors']['red'], img_data['colors']['green'], img_data['colors']['blue'], img_data[
                    'img'] = nce.get_rgb_from_poly(xy_vertices=img_data['xy_vert'])

            return df, img_data

        # This part of the code should only run in case the image was skipped by the touch_test().
        self.log.info('WARNING: DATAFRAME SKIPPED! User region of interest does not touch footprint.shp coordinates.')
        df = pd.DataFrame(columns=list(dd.wfr_vld_names.values()))
        return df, img_data

    def build_raw_csvs(self):
        """
        Parse the input arguments and return a path containing the output intermediary files.
        :return: l1_output_path Posixpath
        """
        self.log.info(f'Searching for WFR files inside: {self.INPUT_DIR}')
        self.log.info('Sorting input files by date.')
        self.sorted_file_list = self.build_list_from_subset(input_directory_path=self.INPUT_DIR)
        self.log.info(f'Input files found: {len(self.sorted_file_list)}')
        self.log.info('------')
        self.log.info(f'Generating ancillary data folder: {self.CSV_N1}')
        Path(self.CSV_N1).mkdir(parents=True, exist_ok=True)
        self.log.info(f'Generating report folder: {self.REP}')
        Path(self.REP).mkdir(parents=True, exist_ok=True)
        self.log.info(f'Attempting to extract geometries from: {self.ROI}')
        self.vertices = Utils.roi2vertex(roi=self.ROI, aux_folder_out=self.CSV_N1)

        total = len(self.sorted_file_list)
        t1 = time.perf_counter()
        done_csvs = []
        for n, img in enumerate(self.sorted_file_list):
            percent = int((n * 100) / total)
            figdate = os.path.basename(img).split('____')[1].split('_')[0]
            self.log.info(f'({percent}%) {n + 1} of {total} - {figdate}')
            try:
                band_data, img_data = self.get_s3_data(wfr_img_folder=img, vertices=self.vertices, roi_file=self.ROI)
                f_b_name = os.path.basename(img).split('.')[0]
                out_dir = os.path.join(self.CSV_N1, f_b_name + '.csv')
                self.log.info(f'Saving DF at : {out_dir}')
                band_data.to_csv(out_dir, index=False)
                done_csvs.append(out_dir)
            except FileNotFoundError as e404:
                # If some Band.nc file was missing inside the image, move to the next one.
                self.log.info(f'{e404}')
                self.log.info(f'Skipping: {figdate}')
                continue

        t2 = time.perf_counter()
        outputstr = f'>>> Finished in {round(t2 - t1, 2)} second(s). <<<'
        self.log.info(outputstr)
        return done_csvs

    def build_single_csv(self, multiFileBridge=False):
        """
        Parse the input arguments and return a path containing the output intermediary file.
        :return: l1_output_path Posixpath
        """
        if not multiFileBridge:  # TODO: build_single_csv should be called by build_raw_csvs for code recycling.
            self.log.info(f'Searching for WFR file inside: {self.INPUT_DIR}')
            self.log.info(f'Generating ancillary data folder: {self.CSV_N1}')
            Path(self.CSV_N1).mkdir(parents=True, exist_ok=True)
            self.log.info(f'Attempting to extract geometries from: {self.ROI}')
            self.vertices = Utils.roi2vertex(roi=self.ROI, aux_folder_out=self.CSV_N1)

        # TODO: https://xarray-spatial.org/reference/_autosummary/xrspatial.multispectral.true_color.html
        band_data, img_data = self.get_s3_data(wfr_img_folder=self.INPUT_DIR, vertices=self.vertices)

        # if df is not None:
        f_b_name = os.path.basename(self.INPUT_DIR).split('.')[0]
        out_dir = os.path.join(self.CSV_N1, f_b_name + '.csv')
        self.log.info(f'Saving DF at : {out_dir}')
        band_data.to_csv(out_dir, index=False)
        return band_data, img_data, [out_dir]

    def process_csv_list(self, raw_csv_list, irmin=False, irmax=False, max_aot=False, use_cams=False, do_clustering=True, k_method='M4'):
        """

        :param max_aot:
        :param k_method:
        :param do_clustering:
        :param use_cams:
        :param irmin:
        :param irmax:
        :param raw_csv_list: [List] containing the absolute path to files extracted by self.get_s3_data
        :return:
        """
        tsgen = TsGenerator(parent_log=self.log)
        self.tsg = tsgen

        # GET SERIES SAVE PATH # TODO: refactor
        safe_version = self.VERSION.replace('.', '-')  # Bad idea to save files with dots in the name
        excel_save_path = os.path.join(self.OUTPUT_DIR, f'{self.RNAME}_SEN3R-{safe_version}.xlsx')
        report_save_path = os.path.join(self.OUTPUT_DIR, f'{self.RNAME}_SEN3R-{safe_version}.pdf')
        out_dir = os.path.join(self.OUTPUT_DIR, 'CSV_N2')
        img_dir = os.path.join(self.OUTPUT_DIR, 'IMG')

        # CREATE THE DIRECTORIES IF THEY DOESN'T EXIST YET
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        Path(img_dir).mkdir(parents=True, exist_ok=True)

        # Start timer
        t1 = time.perf_counter()

        # Update RAW DFs
        total = len(raw_csv_list)

        # List of report pages
        img_report_list = []

        if use_cams:
            # READ CAMS input .csv file
            df_cams = pd.read_csv(self.arguments['cams'])
            df_cams['pydate'] = pd.to_datetime(df_cams['Datetime'])

        for n, img in enumerate(raw_csv_list):

            print(f'>>> Processing: {n + 1} of {total} ... {img}')
            self.log.info(f'>>> Processing: {n + 1} of {total} ... {img}')

            figdate = os.path.basename(img).split('____')[1].split('_')[0]
            figtitl = os.path.basename(out_dir) + '_' + figdate
            savpt_raw_sctr = os.path.join(img_dir, figdate + '_0.png')
            savpt_sctr = os.path.join(img_dir, figdate + '_1.png')
            savpt_rrs = os.path.join(img_dir, figdate + '_2.png')
            savpt_k = os.path.join(img_dir, figdate + '_3.png')

            if use_cams:
                # Find the equivalent observation day in CAMS
                dtlbl = datetime.strptime(figdate, '%Y%m%dT%H%M%S')
                dtlbl = dtlbl.replace(hour=12, minute=0, second=0, microsecond=0)
                cams_row = df_cams[df_cams['pydate'] == dtlbl]
                if len(cams_row) < 1:  # if cams_row == 0 no matching date was found in the CAMS.csv file
                    cams_val = False
                else:
                    cams_val = cams_row['AOD865'].values[0]

            else:
                cams_val = False

            # read LV1 CSVs and generate scatter plots
            rawDf = pd.read_csv(img, sep=',')

            tsgen.plot_sidebyside_sktr(x1_data=rawDf['Oa08_reflectance:float'],
                                       y1_data=rawDf['Oa17_reflectance:float'],
                                       x2_data=rawDf['Oa08_reflectance:float'],
                                       y2_data=rawDf['Oa17_reflectance:float'],
                                       x_lbl='RED: Oa08 (665nm)',
                                       y_lbl='NIR: Oa17 (865nm)',
                                       c1_data=rawDf['A865:float'],
                                       c1_lbl='Aer. Angstrom Expoent (A865)',
                                       c2_data=rawDf['T865:float'],
                                       c2_lbl='Aer. Optical Thickness (T865)',
                                       title=f'RAW {os.path.basename(out_dir)} WFR {figdate} RED:Oa08(665nm) x NIR:Oa17(865nm)',
                                       savepathname=savpt_raw_sctr)

            # reprocessing the raw CSVs and removing reflectances above the threshold in IR.
            try:
                dfpth, df = tsgen.update_csvs(csv_path=img,
                                              glint=20.0,
                                              ir_min_threshold=irmin,
                                              ir_max_threshold=irmax,
                                              savepath=out_dir,
                                              max_aot=max_aot,
                                              cams_val=cams_val)

            except Exception as e:
                self.log.info("type error: " + str(e))
                continue

            if len(df) < 1:
                self.log.info(f'Skipping empty CSV: {dfpth}')
                continue

            # ,--------------------------------------------------,
            # | 26/09/2022 - Generate report page for N... image |----------------------------------------------
            # '--------------------------------------------------'
            if len(df) >= 3:
                self.log.info(f'Dataframe for {figdate} >= 3 pixels: Generating page for report.')
                img_report = tsgen.raw_report(full_csv_path=img,
                                              img_id_date=figdate,
                                              raw_df=rawDf,
                                              filtered_df=df,
                                              output_rprt_path=self.REP)

                img_report_list.append(img_report)
            else:
                self.log.info(f'Dataframe for {figdate} < 3 pixels: Page skipped from the PDF report.')

            # ,--------------------,
            # | DBSCAN Clustering  |------------------------------------------------------------------------------------
            # '--------------------'
            if do_clustering:
                # Backup the DF before cleaning it with DBSCAN
                bkpdf = df.copy()

                # Apply DBSCAN
                tsgen.db_scan(df, dd.clustering_methods[k_method])

                # Plot and save the identified clusters
                tsgen.plot_scattercluster(df, col_x='Oa08_reflectance:float', col_y='Oa17_reflectance:float',
                                          col_color='T865:float', title=f'DBSCAN {figdate}', savepath=savpt_k)

                # Delete rows classified as noise:
                indexNames = df[df['cluster'] == -1].index
                df.drop(indexNames, inplace=True)

                if len(df) > 1:
                    clusters = df.groupby(by='cluster').median()
                    k = Utils.find_nearest(clusters['Oa21_reflectance:float'], 0)
                    # Delete rows from the other clusters:
                    indexNames = df[df['cluster'] != k].index
                    df.drop(indexNames, inplace=True)
                    # TODO : test cluster with the smallest T865 value as a primary/secondary rule.

                else:
                    df = bkpdf.copy()

            tsgen.plot_sidebyside_sktr(x1_data=df['Oa08_reflectance:float'],
                                       y1_data=df['Oa17_reflectance:float'],
                                       x2_data=df['Oa08_reflectance:float'],
                                       y2_data=df['Oa17_reflectance:float'],
                                       x_lbl='RED: Oa08 (665nm)',
                                       y_lbl='NIR: Oa17 (865nm)',
                                       c1_data=df['A865:float'],
                                       c1_lbl='Aer. Angstrom Expoent (A865)',
                                       c2_data=df['T865:float'],
                                       c2_lbl='Aer. Optical Thickness (T865)',
                                       title=f'{os.path.basename(out_dir)} WFR {figdate} RED:Oa08(665nm) x NIR:Oa17(865nm)',
                                       savepathname=savpt_sctr)

            tsgen.s3l2_custom_reflectance_plot(df=df,
                                               figure_title=f'{figdate}\n',
                                               c_lbl='Aer. Optical Thickness (T865)',
                                               save_title=savpt_rrs)

        print(f'Generating EXCEL output at: {excel_save_path}')
        self.log.info(f'Generating EXCEL output at: {excel_save_path}')

        # Generating excel file from the post-processed data
        wdir = out_dir
        todo = tsgen.build_list_from_subset(wdir)

        # Converting and saving the list of mean values into a XLS excel file.
        data = tsgen.generate_tms_data(wdir, todo)

        series_df = pd.DataFrame(data=data)
        # Delete these row indexes from dataFrame
        # indexNames = series_df[series_df['B17-865'] > irmax].index
        # indexNames = series_df[series_df['B17-865'] < irmin].index
        # series_df.drop(indexNames, inplace=True)

        # Compute the avg. SPM and remove decimal precision converting from FLOAT to INT
        series_df['SPM.avg'] = tsgen.get_spm(band865=series_df['B17-865'], band665=series_df['B8-665'])
        #series_df['SPM.avg'] = series_df['SPM.avg'].astype(int)

        # create empty excel
        wb = openpyxl.Workbook()
        wb.save(excel_save_path)

        # open the empty file and fill it up
        book = openpyxl.load_workbook(excel_save_path)
        writer = pd.ExcelWriter(excel_save_path, engine='openpyxl')
        writer.book = book

        # Saving to Excel .xlsx
        series_df.to_excel(writer, sheet_name='wfr', index=False)
        writer.save()
        writer.close()

        # Custom paiting the cells
        # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/colors.html
        wb = openpyxl.load_workbook(excel_save_path)

        # Delete the empty sheet
        del wb['Sheet']

        # Get the sheet containing the final output
        ws = wb['wfr']

        # Create a dictionary of column names
        # https://stackoverflow.com/questions/51478413/select-a-column-by-its-name-openpyxl/51581162
        ColNames = {}
        Current = 0
        for COL in ws.iter_cols(1, ws.max_column):
            ColNames[COL[0].value] = Current
            Current += 1

        quality_column_position = ColNames['Quality']

        mod3r_colors = {0: '00FFFFFF',
                        1: '00008000',
                        2: '00FE6000',
                        3: '00FF0000'}

        for row in ws.iter_rows(min_row=2, min_col=None, max_col=None):
            # get the quality flag for the given row
            flag_qlt = row[quality_column_position]
            for cell in row:
                color_code = mod3r_colors[flag_qlt.value]
                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

        wb.save(excel_save_path)

        # ,------------------------------------,
        # | 26/09/2022 - Generate final report |------------------------------------------------------------------------
        # '------------------------------------'
        self.log.info(f'Generating PDF report at: {report_save_path}')

        img_report_list[0].save(report_save_path, "PDF", resolution=100.0,
                                save_all=True, append_images=img_report_list[1:])

        t2 = time.perf_counter()
        outputstr = f'>>> Finished in {round(t2 - t1, 2)} second(s). <<<'
        print(outputstr)
        self.log.info(outputstr)
